"""
qm_market_news_excel.py — Append Market News Summary sheet to an existing QM Excel workbook.

Inputs:
  --excel <path>             existing .xlsx (e.g., QM_Screener_<date>.xlsx)
  --analyst-json <path>      output of qm_market_news.py
  --news-json <path>         JSON array of {ticker, company, summary} from agent research

Adds a "Market News Summary" sheet with columns:
  A Ticker | B Stock Name | C Current Price | D Analyst Target (Mean) | E Upside % |
  F Consensus | G Market News Summary
"""

import argparse
import json
import sys
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
ALT_ROW_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name="Arial", size=10, italic=True, color="666666")
DATA_FONT = Font(name="Arial", size=10)
BOLD_DATA_FONT = Font(name="Arial", size=10, bold=True)
GREEN_FONT = Font(name="Arial", size=10, color="006100", bold=True)
RED_FONT = Font(name="Arial", size=10, color="C00000", bold=True)


def build_sheet(excel_path, analyst_json_path, news_json_path):
    with open(analyst_json_path, "r") as f:
        analyst = json.load(f)
    with open(news_json_path, "r") as f:
        news = json.load(f)

    by_ticker_analyst = {row["ticker"]: row for row in analyst}
    by_ticker_news = {row["ticker"]: row for row in news}

    wb = load_workbook(excel_path)
    if "Market News Summary" in wb.sheetnames:
        del wb["Market News Summary"]
    ws = wb.create_sheet("Market News Summary")

    columns = [
        ("Ticker", 10),
        ("Stock Name", 26),
        ("Current Price", 14),
        ("Analyst Target (Mean)", 18),
        ("Upside %", 12),
        ("Consensus", 14),
        ("Market News Summary", 110),
    ]

    last_col = get_column_letter(len(columns))
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"].value = "Market News Summary — Strong Buy Stocks (Last 6 Months)"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"].value = (
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | "
        f"Coverage: earnings, analyst actions, product/business, leadership"
    )
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    header_row = 4
    for col_idx, (col_name, width) in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    tickers_ordered = [row["ticker"] for row in news]

    for row_idx, ticker in enumerate(tickers_ordered):
        excel_row = header_row + 1 + row_idx
        is_alt = row_idx % 2 == 1
        ws.row_dimensions[excel_row].height = 90

        a = by_ticker_analyst.get(ticker, {})
        n = by_ticker_news.get(ticker, {})

        company = n.get("company") or a.get("company") or ""
        cur_price = a.get("current_price")
        target = a.get("target_mean")
        upside = a.get("upside_pct")
        rec = a.get("recommendation") or "N/A"
        n_an = a.get("n_analysts") or 0
        consensus = f"{rec.title()} ({n_an} analysts)" if rec and rec != "N/A" else "N/A"
        summary = n.get("summary", "(no summary available)")

        row_values = [ticker, company, cur_price, target, upside, consensus, summary]

        for col_idx, value in enumerate(row_values, 1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if is_alt:
                cell.fill = ALT_ROW_FILL

            col_name = columns[col_idx - 1][0]
            if col_name == "Ticker":
                cell.font = BOLD_DATA_FONT
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_name == "Current Price":
                cell.number_format = '"$"#,##0.00'
                cell.font = BOLD_DATA_FONT
                cell.alignment = Alignment(horizontal="right", vertical="top")
            elif col_name == "Analyst Target (Mean)":
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="top")
            elif col_name == "Upside %":
                cell.number_format = '0.0"%"'
                cell.alignment = Alignment(horizontal="right", vertical="top")
                if isinstance(value, (int, float)):
                    cell.font = GREEN_FONT if value >= 0 else RED_FONT

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row + len(tickers_ordered)}"

    wb.save(excel_path)
    print(f"Appended Market News Summary sheet to {excel_path}", file=sys.stderr)
    print(f"Rows: {len(tickers_ordered)}", file=sys.stderr)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    parser.add_argument("--analyst-json", required=True)
    parser.add_argument("--news-json", required=True)
    args = parser.parse_args()
    build_sheet(args.excel, args.analyst_json, args.news_json)
