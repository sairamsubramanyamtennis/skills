"""
qm_excel.py — Excel report generator for the Quantitative Momentum screener & backtester.

Creates a formatted .xlsx workbook with:
  - Tab 1: "QM Screener" — full ranking with conditional formatting, signals
  - Tab 2: "Backtest Summary" — performance metrics, monthly returns, equity curve data
  - Tab 3: "Market News Summary" — added by qm_market_news.py if requested

Formatting:
  - Dark blue (#1F4E79) headers with white bold text
  - Alternating row fills (#F2F7FB)
  - Signal-based conditional coloring
  - Freeze panes, auto-filter, thin borders
"""

import argparse
import json
import sys
import os
import warnings
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")


# ============================================================================
# STYLE CONSTANTS
# ============================================================================

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
ALT_ROW_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name="Arial", size=10, italic=True, color="666666")
DATA_FONT = Font(name="Arial", size=10)
BOLD_DATA_FONT = Font(name="Arial", size=10, bold=True)

# Signal colors
SIGNAL_COLORS = {
    "Strong Buy": PatternFill(start_color="006100", end_color="006100", fill_type="solid"),
    "Buy":        PatternFill(start_color="548235", end_color="548235", fill_type="solid"),
    "Hold":       PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid"),
    "Sell":       PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
}
SIGNAL_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")


# ============================================================================
# SCREENER SHEET
# ============================================================================

def add_screener_sheet(wb, screener_csv_path):
    """Add the QM Screener ranking sheet to the workbook."""
    df = pd.read_csv(screener_csv_path)
    ws = wb.active
    ws.title = "QM Screener"

    # Title rows
    ws.merge_cells("A1:O1")
    ws["A1"].value = f"Quantitative Momentum Screener — S&P 500"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:O2")
    ws["A2"].value = (
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | "
        f"Methodology: Wes Gray / Alpha Architect (6-1 Momentum + FIP Quality)"
    )
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Blank row 3
    header_row = 4

    # Column definitions
    columns = [
        ("Rank", 8),
        ("Ticker", 10),
        ("Company", 26),
        ("Sector", 22),
        ("6-1M VolAdj", 14),
        ("3M Ret%", 12),
        ("1M Ret%", 12),
        ("FIP Score", 12),
        ("Pct Positive Days", 16),
        ("R-Squared", 12),
        ("Mom Rank", 10),
        ("FIP Rank", 10),
        ("Composite QM Rank", 16),
        ("Percentile", 12),
        ("Signal", 14),
    ]

    # Write headers
    for col_idx, (col_name, width) in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Write data
    for row_idx, (_, row) in enumerate(df.iterrows()):
        excel_row = header_row + 1 + row_idx
        is_alt = row_idx % 2 == 1

        for col_idx, (col_name, _) in enumerate(columns, 1):
            value = row.get(col_name, "")
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

            if is_alt:
                cell.fill = ALT_ROW_FILL

            # Number formatting
            if col_name in ("6-1M VolAdj", "3M Ret%", "1M Ret%"):
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
                # Color: green positive, red negative
                try:
                    if float(value) >= 0:
                        cell.font = Font(name="Arial", size=10, color="006100")
                    else:
                        cell.font = Font(name="Arial", size=10, color="C00000")
                except (ValueError, TypeError):
                    pass
            elif col_name in ("FIP Score", "Pct Positive Days", "R-Squared"):
                cell.number_format = "0.0000"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_name == "Percentile":
                cell.number_format = "0.0"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name in ("Rank", "Mom Rank", "FIP Rank", "Composite QM Rank"):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name == "Ticker":
                cell.font = BOLD_DATA_FONT
            elif col_name == "Signal":
                signal = str(value)
                if signal in SIGNAL_COLORS:
                    cell.fill = SIGNAL_COLORS[signal]
                    cell.font = SIGNAL_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Freeze panes below header
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{header_row + len(df)}"

    return ws


# ============================================================================
# BACKTEST SHEET
# ============================================================================

def add_backtest_sheet(wb, backtest_json_path):
    """Add the Backtest Summary sheet to the workbook."""
    with open(backtest_json_path, "r") as f:
        data = json.load(f)

    ws = wb.create_sheet("Backtest Summary")

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"].value = "Quantitative Momentum — Backtest Results"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:F2")
    ws["A2"].value = (
        f"Period: {data.get('n_years', 'N/A')} years | "
        f"Holdings: {data.get('params', {}).get('n_holdings', 'N/A')} | "
        f"Rebalance: {data.get('params', {}).get('rebalance_freq', 'N/A')}"
    )
    ws["A2"].font = SUBTITLE_FONT

    # Summary metrics
    metrics_start = 4
    metrics = [
        ("Metric", "Value"),
        ("Sharpe Ratio", data.get("sharpe_ratio", "N/A")),
        ("Annual Return", f"{data.get('annual_return', 'N/A')}%"),
        ("Annual Volatility", f"{data.get('annual_vol', 'N/A')}%"),
        ("Max Drawdown", f"{data.get('max_drawdown', 'N/A')}%"),
        ("Total Return", f"{data.get('total_return', 'N/A')}%"),
        ("Excess Return vs SPY", f"{data.get('excess_return', 'N/A')}%"),
        ("SPY Annual Return", f"{data.get('benchmark_annual', 'N/A')}%"),
        ("Months Tested", data.get("n_months", "N/A")),
        ("Rebalances", data.get("n_rebalances", "N/A")),
    ]

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18

    for i, (metric, value) in enumerate(metrics):
        row = metrics_start + i
        cell_a = ws.cell(row=row, column=1, value=metric)
        cell_b = ws.cell(row=row, column=2, value=value)

        if i == 0:
            cell_a.fill = HEADER_FILL
            cell_a.font = HEADER_FONT
            cell_b.fill = HEADER_FILL
            cell_b.font = HEADER_FONT
        else:
            cell_a.font = BOLD_DATA_FONT
            cell_b.font = DATA_FONT
            if i % 2 == 0:
                cell_a.fill = ALT_ROW_FILL
                cell_b.fill = ALT_ROW_FILL

        cell_a.border = THIN_BORDER
        cell_b.border = THIN_BORDER

    # Monthly returns table
    monthly_start = metrics_start + len(metrics) + 2
    ws.cell(row=monthly_start, column=1, value="Monthly Returns").font = BOLD_DATA_FONT

    monthly_rets = data.get("monthly_returns", [])
    if monthly_rets:
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14

        ws.cell(row=monthly_start + 1, column=4, value="Date").fill = HEADER_FILL
        ws.cell(row=monthly_start + 1, column=4).font = HEADER_FONT
        ws.cell(row=monthly_start + 1, column=4).border = THIN_BORDER
        ws.cell(row=monthly_start + 1, column=5, value="Return %").fill = HEADER_FILL
        ws.cell(row=monthly_start + 1, column=5).font = HEADER_FONT
        ws.cell(row=monthly_start + 1, column=5).border = THIN_BORDER

        for j, mr in enumerate(monthly_rets):
            row = monthly_start + 2 + j
            ws.cell(row=row, column=4, value=mr["date"]).font = DATA_FONT
            ws.cell(row=row, column=4).border = THIN_BORDER
            ret_cell = ws.cell(row=row, column=5, value=round(mr["return"] * 100, 2))
            ret_cell.number_format = "0.00"
            ret_cell.font = DATA_FONT
            ret_cell.border = THIN_BORDER
            if j % 2 == 1:
                ws.cell(row=row, column=4).fill = ALT_ROW_FILL
                ret_cell.fill = ALT_ROW_FILL

    # Parameters used
    params_start = metrics_start + len(metrics) + 2
    ws.cell(row=params_start, column=1, value="Parameters Used").font = BOLD_DATA_FONT
    params = data.get("params", {})
    for i, (k, v) in enumerate(params.items()):
        row = params_start + 1 + i
        ws.cell(row=row, column=1, value=k).font = DATA_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=str(v)).font = DATA_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER

    ws.freeze_panes = "A4"
    return ws


# ============================================================================
# MAIN
# ============================================================================

def generate_report(screener_csv=None, backtest_json=None, output_path="qm_report.xlsx"):
    """Generate the Excel report."""
    wb = Workbook()

    if screener_csv and os.path.exists(screener_csv):
        add_screener_sheet(wb, screener_csv)
        print(f"Added QM Screener sheet ({screener_csv})", file=sys.stderr)
    else:
        # Remove default sheet if no screener data
        ws = wb.active
        ws.title = "Info"
        ws["A1"] = "No screener data provided"

    if backtest_json and os.path.exists(backtest_json):
        add_backtest_sheet(wb, backtest_json)
        print(f"Added Backtest Summary sheet ({backtest_json})", file=sys.stderr)

    wb.save(output_path)
    print(f"Saved report: {output_path}", file=sys.stderr)
    return output_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="QM Excel Report Generator")
    parser.add_argument("--screener-data", type=str, default=None,
                        help="Path to qm_full_ranking.csv")
    parser.add_argument("--backtest-data", type=str, default=None,
                        help="Path to qm_backtest_results.json")
    parser.add_argument("--output", type=str, default="qm_report.xlsx",
                        help="Output Excel file path")
    args = parser.parse_args()

    generate_report(args.screener_data, args.backtest_data, args.output)
